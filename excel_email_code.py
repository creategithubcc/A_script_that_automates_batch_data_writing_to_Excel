import os
import re
import openpyxl
from openpyxl.utils import get_column_letter

# 创建Excel工作簿
workbook = openpyxl.Workbook()
sheet = workbook.active

# 添加标题
titles = ['a', 'b', 'c', 'd', 'e']#行
sheet['A1'] = 'name'#列
for i, title in enumerate(titles):
    column_letter = get_column_letter(i + 2)
    sheet[f'{column_letter}1'] = title

# 读取EML文件夹中的邮件
eml_folder = 'file'
row = 2  # 从第二行开始填写数据
for filename in os.listdir(eml_folder):
    if filename.endswith('.eml'):
        with open(os.path.join(eml_folder, filename), 'r',encoding='utf-8') as eml_file:
            email_content = eml_file.read()

        # 提取收件人
        recipient_match = re.findall(r'Delivered-To: (\S+)', email_content)
        if recipient_match:
            for recipient in recipient_match:
                # 截取邮件地址的一半字符串
                recipient_half = recipient[:len(recipient) // 2]
                sheet[f'A{row}'] = recipient_half

                # 检查邮件地址的前缀并填写相应的列
                for i, title in enumerate(titles):
                    column_letter = get_column_letter(i + 2)
                    if recipient.startswith(title):
                        sheet[f'{column_letter}{row}'] = 1
                    else:
                        sheet[f'{column_letter}{row}'] = 0
                row += 1

# 保存Excel文件
workbook.save('email_analysis.xlsx')
