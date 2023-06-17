import os
import re
import openpyxl
from openpyxl.utils import get_column_letter

# 创建Excel工作簿
workbook = openpyxl.Workbook()
sheet = workbook.active

# 添加标题
titles = ['a', 'b', 'c', 'd', 'e']#行
sheet['A1'] = 'name'#列
for i, title in enumerate(titles):
    column_letter = get_column_letter(i + 2)
    sheet[f'{column_letter}1'] = title

# 读取EML文件夹中的邮件
eml_folder = 'file'#文件夹
recipient_dict = {}
for filename in os.listdir(eml_folder):
    if filename.endswith('.eml'):#读取什么类型的文件
        with open(os.path.join(eml_folder, filename), 'r',encoding='utf-8') as eml_file:
            email_content = eml_file.read()

        # 提取收件人
        recipient_match = re.findall(r'Delivered-To: (\S+)', email_content)#找到里面的关键词数据
        if recipient_match:
            for recipient in recipient_match:
                recipient_half = recipient[:len(recipient) // 2]#取对半
                recipient_domain = recipient_half.split('@')[1]

                # 检查邮件地址的前缀并更新相应的列
                if recipient_domain not in recipient_dict:
                    recipient_dict[recipient_domain] = [0] * len(titles)

                for i, title in enumerate(titles):
                    if recipient.startswith(title):
                        recipient_dict[recipient_domain][i] = 1

# 填写Excel表格
row = 2  # 从第二行开始填写数据
for domain, columns in recipient_dict.items():
    row_values = [domain] + columns
    sheet.append(row_values)
    row += 1

# 合并相同域名的行
for row in range(2, sheet.max_row + 1):
    for col in range(2, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value == 1:
            for merge_row in range(row + 1, sheet.max_row + 1):
                if sheet.cell(row=merge_row, column=1).value == sheet.cell(row=row, column=1).value:
                    sheet.cell(row=merge_row, column=col).value = 1
                else:
                    break

# 删除重复的行
for row in range(sheet.max_row, 2, -1):
    if sheet.cell(row=row, column=1).value == sheet.cell(row=row - 1, column=1).value:
        sheet.delete_rows(row)

# 保存Excel文件
workbook.save('new_email_analysis.xlsx')
